#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import os
import queue
import re
import sys
import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable, List, Optional, Set, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from PySide6.QtCore import QTimer
    from PySide6.QtWidgets import (
        QApplication,
        QCheckBox,
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QProgressBar,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
    PYSIDE_AVAILABLE = True
except Exception:
    PYSIDE_AVAILABLE = False


URL_SPLIT_RE = re.compile(r"\s*[|,，;\n\r]+\s*")
IMAGE_EXTS = [".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tif", ".tiff"]


def sanitize_path_component(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[\\/:*?\"<>|\n\r\t]+", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or "未知"


def normalize_url(url: str) -> str:
    u = (url or "").strip()
    if not u:
        return ""
    u = u.replace("\u3000", " ").strip()
    u = u.replace("\\", "/")
    if u.startswith("https:/") and not u.startswith("https://"):
        u = "https://" + u[len("https:/") :].lstrip("/")
    if u.startswith("http:/") and not u.startswith("http://"):
        u = "http://" + u[len("http:/") :].lstrip("/")
    return u


def split_urls(cell_value) -> List[str]:
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    parts = URL_SPLIT_RE.split(s)
    urls: List[str] = []
    for p in parts:
        u = normalize_url(p)
        if u:
            urls.append(u)
    return urls


def is_http_url(url: str) -> bool:
    u = (url or "").strip().lower()
    return u.startswith("http://") or u.startswith("https://")


def guess_ext_from_url(url: str) -> str:
    try:
        path = requests.utils.urlparse(url).path
    except Exception:
        return ""
    ext = os.path.splitext(path)[1].lower()
    if ext in IMAGE_EXTS:
        return ext
    return ""


def guess_ext_from_content_type(content_type: str) -> str:
    ct = (content_type or "").split(";")[0].strip().lower()
    mapping = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/gif": ".gif",
        "image/webp": ".webp",
        "image/bmp": ".bmp",
        "image/tiff": ".tif",
    }
    return mapping.get(ct, "")


@dataclass
class DownloadResult:
    ok: bool
    path: Optional[Path]
    error: Optional[str]


@dataclass
class DownloadTask:
    row: int
    name: str
    phone: str
    label: str
    index: int
    url: str
    dest_path_no_ext: Path


@dataclass
class DownloadSummary:
    total: int = 0
    ok: int = 0
    failed: int = 0
    failed_rows: Set[int] = field(default_factory=set)
    cancelled: bool = False


def download_one(
    session: requests.Session,
    url: str,
    dest_path_no_ext: Path,
    *,
    timeout_s: int,
    retries: int,
    backoff_s: float,
    skip_existing: bool,
    should_stop: Optional[Callable[[], bool]] = None,
) -> DownloadResult:
    if should_stop and should_stop():
        return DownloadResult(False, None, "用户已停止")
    if skip_existing:
        for ext in IMAGE_EXTS:
            p = dest_path_no_ext.with_suffix(ext)
            if p.exists() and p.stat().st_size > 0:
                return DownloadResult(True, p, None)

    last_err = None
    for attempt in range(retries + 1):
        if should_stop and should_stop():
            return DownloadResult(False, None, "用户已停止")
        try:
            with session.get(url, stream=True, timeout=timeout_s, allow_redirects=True) as resp:
                resp.raise_for_status()

                ext = guess_ext_from_url(url) or guess_ext_from_content_type(resp.headers.get("Content-Type", ""))
                if not ext:
                    ext = ".bin"
                dest_path = dest_path_no_ext.with_suffix(ext)
                dest_path.parent.mkdir(parents=True, exist_ok=True)

                tmp_path = dest_path.with_suffix(dest_path.suffix + ".part")
                with open(tmp_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=1024 * 256):
                        if should_stop and should_stop():
                            tmp_path.unlink(missing_ok=True)
                            return DownloadResult(False, None, "用户已停止")
                        if chunk:
                            f.write(chunk)

                if tmp_path.stat().st_size == 0:
                    tmp_path.unlink(missing_ok=True)
                    raise RuntimeError("下载结果为空文件")

                tmp_path.replace(dest_path)
                return DownloadResult(True, dest_path, None)
        except Exception as e:
            last_err = str(e)
            if attempt < retries:
                time.sleep(backoff_s * (2**attempt))
            continue

    return DownloadResult(False, None, last_err or "未知错误")


def iter_rows(ws, start_row: int) -> Iterable[int]:
    for r in range(start_row, ws.max_row + 1):
        yield r


def build_tasks(xlsx_path: Path, sheet_name: Optional[str], out_root: Path, start_row: int) -> List[DownloadTask]:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        col_name = 2
        col_phone = 3
        all_cols = list(range(1, ws.max_column + 1))
        col_labels: List[Tuple[int, str]] = []
        for c in all_cols:
            label = ws.cell(row=2, column=c).value
            label_s = sanitize_path_component(str(label)) if label not in (None, "") else get_column_letter(c)
            col_labels.append((c, label_s))

        tasks: List[DownloadTask] = []
        for r in iter_rows(ws, start_row):
            name = sanitize_path_component(str(ws.cell(r, col_name).value or ""))
            phone = sanitize_path_component(str(ws.cell(r, col_phone).value or ""))
            if name == "未知" and phone == "未知":
                continue

            person_dir = out_root / f"{name}_{phone}"
            for c, label in col_labels:
                urls = split_urls(ws.cell(r, c).value)
                valid_urls = [u for u in urls if is_http_url(u)]
                for i, url in enumerate(valid_urls, start=1):
                    tasks.append(
                        DownloadTask(
                            row=r,
                            name=name,
                            phone=phone,
                            label=label,
                            index=i,
                            url=url,
                            dest_path_no_ext=person_dir / f"{label}_{i:02d}",
                        )
                    )
        return tasks
    finally:
        wb.close()


def run_download(
    *,
    xlsx_path: Path,
    sheet_name: Optional[str],
    out_root: Path,
    start_row: int,
    timeout_s: int,
    retries: int,
    backoff_s: float,
    skip_existing: bool,
    use_proxy: bool,
    progress_cb: Optional[Callable[[int, int, str], None]] = None,
    log_cb: Optional[Callable[[str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
) -> DownloadSummary:
    tasks = build_tasks(xlsx_path, sheet_name, out_root, start_row)
    total = len(tasks)
    summary = DownloadSummary(total=total)
    if total == 0:
        return summary

    session = requests.Session()
    session.trust_env = bool(use_proxy)
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (compatible; download_excel_images/2.0)",
            "Accept": "image/avif,image/webp,image/*,*/*;q=0.8",
        }
    )

    done = 0
    for task in tasks:
        if should_stop and should_stop():
            summary.cancelled = True
            break
        res = download_one(
            session,
            task.url,
            task.dest_path_no_ext,
            timeout_s=timeout_s,
            retries=retries,
            backoff_s=backoff_s,
            skip_existing=skip_existing,
            should_stop=should_stop,
        )
        if should_stop and should_stop() and not res.ok and res.error == "用户已停止":
            summary.cancelled = True
            break
        done += 1
        if res.ok:
            summary.ok += 1
            msg = f"[OK] 行{task.row} {task.name}_{task.phone} {task.label}#{task.index} -> {res.path.name}"
        else:
            summary.failed += 1
            summary.failed_rows.add(task.row)
            msg = f"[FAIL] 行{task.row} {task.name}_{task.phone} {task.label}#{task.index} {task.url} | {res.error}"

        if log_cb:
            log_cb(msg)
        if progress_cb:
            progress_cb(done, total, msg)

    return summary


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="从 Excel 的 L-T 单元格下载图片链接。")
    p.add_argument("--xlsx", default="摩托车消费券核销记录.xlsx", help="Excel 文件路径")
    p.add_argument("--sheet", default=None, help="工作表名称（默认第一个）")
    p.add_argument("--out", default="下载图片", help="输出根目录")
    p.add_argument("--start-row", type=int, default=3, help="开始处理的行号")
    p.add_argument("--timeout", type=int, default=30, help="单个请求超时秒数")
    p.add_argument("--retries", type=int, default=2, help="失败重试次数")
    p.add_argument("--backoff", type=float, default=0.8, help="重试退避基数秒")
    p.add_argument("--skip-existing", action="store_true", help="如目标文件已存在则跳过")
    p.add_argument("--use-proxy", action="store_true", help="使用系统代理环境变量")
    p.add_argument("--gui", action="store_true", help="强制启动图形界面")
    return p.parse_args(argv)


class AppWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Excel 图片下载器 (PySide6)")
        self.resize(900, 620)

        self.running = False
        self.stop_event = threading.Event()
        self.resume_checked = True
        self.event_queue: "queue.Queue[Tuple[str, object]]" = queue.Queue()

        self._build_ui()
        self.timer = QTimer(self)
        self.timer.timeout.connect(self._drain_events)
        self.timer.start(100)

    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)
        root = QVBoxLayout(central)

        line1 = QHBoxLayout()
        line1.addWidget(QLabel("Excel 文件:"))
        self.xlsx_edit = QLineEdit()
        line1.addWidget(self.xlsx_edit, 1)
        choose_excel_btn = QPushButton("选择文件")
        choose_excel_btn.clicked.connect(self._choose_excel)
        line1.addWidget(choose_excel_btn)
        root.addLayout(line1)

        line2 = QHBoxLayout()
        line2.addWidget(QLabel("输出目录:"))
        self.out_edit = QLineEdit(str((Path.cwd() / "下载图片").resolve()))
        line2.addWidget(self.out_edit, 1)
        choose_out_btn = QPushButton("选择目录")
        choose_out_btn.clicked.connect(self._choose_out_dir)
        line2.addWidget(choose_out_btn)
        root.addLayout(line2)

        line3 = QHBoxLayout()
        self.start_btn = QPushButton("开始下载")
        self.start_btn.clicked.connect(self._start_download)
        line3.addWidget(self.start_btn)
        self.stop_btn = QPushButton("停止下载")
        self.stop_btn.clicked.connect(self._stop_download)
        self.stop_btn.setEnabled(False)
        line3.addWidget(self.stop_btn)
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        line3.addWidget(self.progress, 1)
        root.addLayout(line3)

        line4 = QHBoxLayout()
        self.resume_checkbox = QCheckBox("继续下载（跳过已下载）")
        self.resume_checkbox.setChecked(True)
        self.resume_checkbox.stateChanged.connect(self._on_resume_changed)
        line4.addWidget(self.resume_checkbox)
        line4.addStretch(1)
        root.addLayout(line4)

        self.status_label = QLabel("请选择 Excel 文件后开始下载")
        root.addWidget(self.status_label)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        root.addWidget(self.log_text, 1)

    def _choose_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 Excel 文件",
            "",
            "Excel 文件 (*.xlsx *.xlsm *.xltx *.xltm);;所有文件 (*)",
        )
        if path:
            self.xlsx_edit.setText(path)

    def _choose_out_dir(self) -> None:
        path = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if path:
            self.out_edit.setText(path)

    def _append_log(self, text: str) -> None:
        self.log_text.append(text)

    def _on_resume_changed(self) -> None:
        self.resume_checked = self.resume_checkbox.isChecked()

    def _start_download(self) -> None:
        if self.running:
            return
        xlsx = self.xlsx_edit.text().strip()
        out = self.out_edit.text().strip()
        if not xlsx:
            QMessageBox.warning(self, "提示", "请先选择 Excel 文件")
            return
        xlsx_path = Path(xlsx).expanduser()
        if not xlsx_path.exists():
            QMessageBox.critical(self, "错误", f"Excel 文件不存在：\n{xlsx_path}")
            return
        out_path = Path(out).expanduser()
        out_path.mkdir(parents=True, exist_ok=True)

        self.running = True
        self.stop_event.clear()
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setValue(0)
        self.log_text.clear()
        self.status_label.setText("准备中...")
        self._append_log(f"开始下载，Excel：{xlsx_path}")
        self._append_log(f"输出目录：{out_path}")
        self._append_log(f"继续下载模式：{'开启' if self.resume_checked else '关闭'}")

        worker = threading.Thread(
            target=self._worker_download,
            args=(xlsx_path.resolve(), out_path.resolve()),
            daemon=True,
        )
        worker.start()

    def _stop_download(self) -> None:
        if not self.running:
            return
        self.stop_event.set()
        self.status_label.setText("正在停止下载...")
        self._append_log("收到停止指令，正在安全停止...")
        self.stop_btn.setEnabled(False)

    def _worker_download(self, xlsx_path: Path, out_path: Path) -> None:
        try:
            summary = run_download(
                xlsx_path=xlsx_path,
                sheet_name=None,
                out_root=out_path,
                start_row=3,
                timeout_s=30,
                retries=2,
                backoff_s=0.8,
                skip_existing=self.resume_checked,
                use_proxy=False,
                progress_cb=lambda done, total, msg: self.event_queue.put(("progress", (done, total, msg))),
                log_cb=lambda msg: self.event_queue.put(("log", msg)),
                should_stop=lambda: self.stop_event.is_set(),
            )
            self.event_queue.put(("done", summary))
        except Exception as e:
            self.event_queue.put(("error", str(e)))

    def _drain_events(self) -> None:
        try:
            while True:
                event, payload = self.event_queue.get_nowait()
                if event == "log":
                    self._append_log(str(payload))
                elif event == "progress":
                    done, total, _msg = payload
                    percent = (done / total * 100.0) if total else 0.0
                    self.progress.setValue(int(percent))
                    self.status_label.setText(f"下载进度：{done}/{total} ({percent:.1f}%)")
                elif event == "done":
                    summary: DownloadSummary = payload
                    self.running = False
                    self.start_btn.setEnabled(True)
                    self.stop_btn.setEnabled(False)
                    self.progress.setValue(100 if summary.total else 0)
                    fail_rows_sorted = sorted(summary.failed_rows)
                    fail_rows_text = "无" if not fail_rows_sorted else ", ".join(str(r) for r in fail_rows_sorted)
                    if summary.cancelled:
                        self.resume_checkbox.setChecked(True)
                        result = (
                            f"下载已停止。\n总数：{summary.total}\n已完成：{summary.ok + summary.failed}\n"
                            f"成功：{summary.ok}\n失败：{summary.failed}\n失败行号：{fail_rows_text}"
                        )
                        self.status_label.setText(
                            f"已停止：总数 {summary.total}，已完成 {summary.ok + summary.failed}"
                        )
                        title = "下载已停止"
                    else:
                        result = (
                            f"下载结束。\n总数：{summary.total}\n成功：{summary.ok}\n失败：{summary.failed}\n"
                            f"失败行号：{fail_rows_text}"
                        )
                        self.status_label.setText(
                            f"完成：总数 {summary.total}，成功 {summary.ok}，失败 {summary.failed}"
                        )
                        title = "下载完成"
                    self._append_log(result)
                    QMessageBox.information(self, title, result)
                elif event == "error":
                    self.running = False
                    self.start_btn.setEnabled(True)
                    self.stop_btn.setEnabled(False)
                    self.status_label.setText("执行失败")
                    QMessageBox.critical(self, "错误", str(payload))
        except queue.Empty:
            pass


def run_cli(args: argparse.Namespace) -> int:
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    out_root = Path(args.out).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"找不到 Excel：{xlsx_path}", file=sys.stderr)
        return 2
    out_root.mkdir(parents=True, exist_ok=True)

    summary = run_download(
        xlsx_path=xlsx_path,
        sheet_name=args.sheet,
        out_root=out_root,
        start_row=args.start_row,
        timeout_s=args.timeout,
        retries=args.retries,
        backoff_s=args.backoff,
        skip_existing=args.skip_existing,
        use_proxy=args.use_proxy,
        progress_cb=lambda done, total, _msg: print(f"\r进度：{done}/{total}", end="", flush=True),
        log_cb=lambda msg: print(f"\n{msg}"),
    )
    print()
    failed_rows = sorted(summary.failed_rows)
    print(f"完成：总数 {summary.total}，成功 {summary.ok}，失败 {summary.failed}。输出目录：{out_root}")
    print(f"失败行号：{failed_rows if failed_rows else '无'}")
    return 0 if summary.failed == 0 else 1


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    force_gui = bool(args.gui)
    no_cli_args = len(sys.argv) == 1 if argv is None else len(argv) == 0
    if force_gui and not PYSIDE_AVAILABLE:
        print("当前 Python 环境缺少 PySide6，无法启动图形界面。", file=sys.stderr)
        print("请先安装：pip install pyside6", file=sys.stderr)
        return 2
    if (force_gui or no_cli_args) and PYSIDE_AVAILABLE:
        app = QApplication(sys.argv if argv is None else ["download_excel_images.py", *argv])
        window = AppWindow()
        window.show()
        app.exec()
        return 0
    if no_cli_args and not PYSIDE_AVAILABLE:
        print("检测到当前环境缺少 PySide6，已自动切换到命令行模式。")
        print("请使用 --xlsx 指定 Excel 文件，例如：")
        print('python3 download_excel_images.py --xlsx "摩托车消费券核销记录.xlsx" --out "下载图片"')
    return run_cli(args)


if __name__ == "__main__":
    raise SystemExit(main())

